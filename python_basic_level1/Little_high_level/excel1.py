from openpyxl import Workbook, load_workbook

w1=Workbook()
w2=w1.active
w2.title="data"

w2.append(["data1","data2"])
w1.save("input.xlsx")

wb = load_workbook("input.xlsx")
ws = wb.active
for row in ws.iter_rows(values_only=True):
    print(row)
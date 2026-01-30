import os
import pandas as pd
from datetime import datetime
from openpyxl import load_workbook

schema_name = 'BDAAS_CONFORMED'
table_name = [ 'TRANSACTION_2022_2025_V5']
date_list = [20220101]

current_path = os.getcwd()
output_base = os.path.join(current_path, "output")

today = datetime.now().strftime("%Y%m%d")
today_folder = os.path.join(output_base, today)
os.makedirs(today_folder, exist_ok=True)

for tab in table_name:
    excel_path = os.path.join(today_folder, f"{tab}.xlsx")
    dq_conn = metadata_connections()

    for dte in date_list:

        # --- Insert execution summary ---
        insert_query = f"""
        insert into ddq.ddq_rules_execution_summary
        (job_name, mapping_id, schema_name, table_name, job_status,
         executed_by, processed_date, executed_on, framework_name)
        values
        ('validation','adhoc','{schema_name}','{tab}',
         'queued','mahesp@deloitte.com',{dte},
         CURRENT_TIMESTAMP,'great_expectations')
        """
        dq_conn.query_execution(query=insert_query)

        # --- Run GE notebook ---
        dbutils.notebook.run(
            "/Users/mahesp_deloitte.com#ext#@opencloud.onmicrosoft.com/DQ_WRAPPER/DQ_GE_Framework_Snowflake/main",
            timeout_seconds=0,
            arguments={"table_name": tab, "processed_date": str(dte)}
        )

        job_id_query = f"""
        select max(job_execution_id) from ddq.ddq_rules_execution_summary
        """
        val = dq_conn.query_fetchone(query=job_id_query)

        # --- Fetch details ---
        detail_query = f"""
        select rule_execution_id, schema_name, table_name, column_name,
               rule_description,  constraint_status,  total_scanned,
               pass_count, failed_count 
               from ddq.DDQ_RULES_EXECUTION_DETAIL
        where job_execution_id={val[0]}
        """
        print("Running fetch all code")
        detail = dq_conn.query_fetch_all(query=detail_query)

        columns = [
            'rule_execution_id','schema_name','table_name','column_name',
            'rule_description','constraint_status', 'total_scanned',
            'pass_count','failed_count'
        ]

        rows = [tuple(row) for row in detail]
        df = pd.DataFrame(rows, columns=columns)

        # ⭐ ADD execution_date column
        df['execution_date'] = dte
        SHEET_NAME = "data"

        if not os.path.exists(excel_path):
            # File does not exist → create
            with pd.ExcelWriter(excel_path, engine='openpyxl', mode='w') as writer:
                df.to_excel(writer, sheet_name=SHEET_NAME, index=False)

        else:
            print("data is")
            # File exists → check if sheet exists
            wb = load_workbook(excel_path)
            if SHEET_NAME in wb.sheetnames:
                existing_df = pd.read_excel(excel_path, sheet_name=SHEET_NAME)
                final_df = pd.concat([existing_df, df], ignore_index=True)
            else:
                # Sheet missing → treat as first write
                final_df = df

            # Always rewrite the Excel safely
            with pd.ExcelWriter(excel_path, engine='openpyxl', mode='w') as writer:
                final_df.to_excel(writer, sheet_name=SHEET_NAME, index=False)

print(f"All files saved in folder: {today_folder}")